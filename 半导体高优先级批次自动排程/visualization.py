"""甘特图和Excel输出模块 —— Lot维度 + 设备维度甘特图 + Excel导出"""
import os
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional
from models import ScheduleEntry, EqpScheduleEntry, QTimeAlert

# CJK 字体注册
import matplotlib.font_manager as fm
_font_path = os.path.expanduser('~/.fonts/NotoSansCJKsc-Regular.otf')
if os.path.exists(_font_path):
    fm.fontManager.addfont(_font_path)
    import matplotlib
    matplotlib.rcParams['font.family'] = 'Noto Sans CJK SC'
    matplotlib.rcParams['axes.unicode_minus'] = False

# 颜色映射
COLORS = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
    "#aec7e8", "#ffbb78", "#98df8a", "#ff9896", "#c5b0d5",
    "#c49c94", "#f7b6d2", "#c7c7c7", "#dbdb8d", "#9edae5",
]


def _short_step_name(step_name: str) -> str:
    """裁剪步骤名前缀 XXXX- 格式，如 A005-P1-FC-DUMMY → P1-FC-DUMMY"""
    if len(step_name) > 5 and step_name[4] == "-":
        return step_name[5:]
    return step_name


def _step_color_map(lot_entries: list[ScheduleEntry]) -> dict[str, str]:
    unique_steps = list(dict.fromkeys(e.step_name for e in lot_entries))
    return {step: COLORS[i % len(COLORS)] for i, step in enumerate(unique_steps)}


def _lot_color_map(eqp_entries: list[EqpScheduleEntry]) -> dict[str, str]:
    unique_lots = list(dict.fromkeys(e.lot_name for e in eqp_entries))
    return {lot: COLORS[i % len(COLORS)] for i, lot in enumerate(unique_lots)}


def _add_shift_lines(ax, t_min, t_max, shift_times: list[tuple[int, int]] = None):
    """在时间轴上添加班次边界虚线"""
    if shift_times is None:
        shift_times = [(8, 30), (20, 30)]
    current = t_min.replace(hour=0, minute=0, second=0, microsecond=0)
    while current <= t_max + timedelta(days=1):
        for h, m in shift_times:
            shift_time = current.replace(hour=h, minute=m, second=0, microsecond=0)
            if t_min <= shift_time <= t_max:
                ax.axvline(x=shift_time, color="gray", linestyle="--", linewidth=0.5, alpha=0.5)
        current += timedelta(days=1)


def plot_lot_gantt(
    lot_entries: list[ScheduleEntry],
    output_path: str,
    time_start: Optional[datetime] = None,
    time_end: Optional[datetime] = None,
    lot_filter: Optional[list[str]] = None,
    shift_times: list[tuple[int, int]] = None,
):
    """生成 Lot 维度甘特图 —— 每行一个Lot，标注步骤名称"""
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import matplotlib.patches as mpatches

    # 筛选
    entries = lot_entries
    if time_start:
        entries = [e for e in entries if e.end_time >= time_start]
    if time_end:
        entries = [e for e in entries if e.start_time <= time_end]
    if lot_filter:
        lot_filter_set = set(lot_filter)
        entries = [e for e in entries if e.lot_name in lot_filter_set]

    if not entries:
        print("无数据，跳过 Lot 甘特图生成")
        return

    step_colors = _step_color_map(entries)
    lot_names = list(dict.fromkeys(e.lot_name for e in entries))

    # 计算时间范围
    t_min = min(e.start_time for e in entries)
    t_max = max(e.end_time for e in entries)
    t_padding = timedelta(hours=2)
    t_min -= t_padding
    t_max += t_padding

    n_lots = len(lot_names)
    # 根据行数动态计算：行高、字体大小
    row_height = max(0.6, min(1.0, 8.0 / max(n_lots, 1)))
    bar_height = row_height * 0.85
    bar_fontsize = max(3.5, min(6.5, row_height * 8))
    ylabel_fontsize = max(7, min(12, row_height * 14))
    fig, ax = plt.subplots(figsize=(24, max(6, n_lots * row_height)))

    for i, lot_name in enumerate(lot_names):
        lot_steps = [e for e in entries if e.lot_name == lot_name]
        # 按开始时间排序
        lot_steps.sort(key=lambda e: e.start_time)
        for step in lot_steps:
            color = step_colors[step.step_name]
            start_num = mdates.date2num(step.start_time)
            end_num = mdates.date2num(step.end_time)
            duration = end_num - start_num
            bar = ax.barh(i, duration, left=start_num, height=bar_height,
                          color=color, edgecolor="white", linewidth=0.3)

            # 标注步骤名称（竖排，宽度足够时显示）
            if duration > 0.015:  # ~20分钟
                mid = start_num + duration / 2
                label = _short_step_name(step.step_name)
                ax.text(mid, i, label, ha="center", va="center",
                        fontsize=bar_fontsize, color="white", fontweight="bold",
                        rotation=90)

    # 班次边界线
    _add_shift_lines(ax, t_min, t_max, shift_times)

    ax.set_yticks(range(n_lots))
    ax.set_yticklabels(lot_names, fontsize=ylabel_fontsize, rotation=90, va="center")
    ax.set_ylim(-0.5, n_lots - 0.5)
    ax.invert_yaxis()
    ax.set_xlim(mdates.date2num(t_min), mdates.date2num(t_max))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d %H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=6))
    plt.xticks(rotation=45, fontsize=8)
    ax.set_xlabel("Time", fontsize=12)
    ax.set_title("Lot Dimension Gantt Chart", fontsize=14, fontweight="bold")
    ax.grid(axis="x", linestyle=":", alpha=0.3)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Lot Gantt chart saved: {output_path}")


def plot_eqp_gantt(
    eqp_entries: list[EqpScheduleEntry],
    output_path: str,
    time_start: Optional[datetime] = None,
    time_end: Optional[datetime] = None,
    eqp_filter: Optional[list[str]] = None,
    shift_times: list[tuple[int, int]] = None,
):
    """生成设备维度甘特图 —— 每行一台设备，标注Lot名称和步骤"""
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import matplotlib.patches as mpatches

    # 筛选
    entries = eqp_entries
    if time_start:
        entries = [e for e in entries if e.end_time >= time_start]
    if time_end:
        entries = [e for e in entries if e.start_time <= time_end]
    if eqp_filter:
        eqp_filter_set = set(eqp_filter)
        entries = [e for e in entries if e.eqp_id in eqp_filter_set]

    if not entries:
        print("无数据，跳过设备甘特图生成")
        return

    lot_colors = _lot_color_map(entries)
    eqp_ids = list(dict.fromkeys(e.eqp_id for e in entries))

    # 计算时间范围
    t_min = min(e.start_time for e in entries)
    t_max = max(e.end_time for e in entries)
    t_padding = timedelta(hours=2)
    t_min -= t_padding
    t_max += t_padding

    n_eqps = len(eqp_ids)
    # 根据行数动态计算：行高、字体大小
    row_height = max(0.55, min(0.9, 8.0 / max(n_eqps, 1)))
    bar_height = row_height * 0.85
    bar_fontsize = max(3.0, min(5.5, row_height * 7))
    ylabel_fontsize = max(7, min(12, row_height * 14))
    fig, ax = plt.subplots(figsize=(24, max(6, n_eqps * row_height)))

    for i, eqp_id in enumerate(eqp_ids):
        eqp_steps = [e for e in entries if e.eqp_id == eqp_id]
        eqp_steps.sort(key=lambda e: e.start_time)
        for step in eqp_steps:
            color = lot_colors[step.lot_name]
            start_num = mdates.date2num(step.start_time)
            end_num = mdates.date2num(step.end_time)
            duration = end_num - start_num
            ax.barh(i, duration, left=start_num, height=bar_height,
                    color=color, edgecolor="white", linewidth=0.3)

            # 标注 Lot 名称和步骤（竖排）
            if duration > 0.015:
                mid = start_num + duration / 2
                short_step = _short_step_name(step.step_name)
                label = f"{step.lot_name}  {short_step}"
                ax.text(mid, i, label, ha="center", va="center",
                        fontsize=bar_fontsize, color="white", fontweight="bold",
                        rotation=90)

    # 班次边界线
    _add_shift_lines(ax, t_min, t_max, shift_times)

    ax.set_yticks(range(n_eqps))
    ax.set_yticklabels(eqp_ids, fontsize=ylabel_fontsize, rotation=0, ha="right")
    ax.set_ylim(-0.5, n_eqps - 0.5)
    ax.invert_yaxis()
    ax.set_xlim(mdates.date2num(t_min), mdates.date2num(t_max))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d %H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=6))
    plt.xticks(rotation=45, fontsize=8)
    ax.set_xlabel("Time", fontsize=12)
    ax.set_title("Equipment Dimension Gantt Chart", fontsize=14, fontweight="bold")
    ax.grid(axis="x", linestyle=":", alpha=0.3)

    # 图例
    legend_patches = [mpatches.Patch(color=lot_colors[l], label=l) for l in lot_colors]
    ax.legend(handles=legend_patches, loc="upper right", fontsize=8, ncol=2)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Eqp Gantt chart saved: {output_path}")


def _parse_shift_starts(shift_configs):
    """解析 shift_configs 为班次开始时间列表 [(shift_name, h, m), ...]"""
    shift_starts = []
    if shift_configs:
        for sc in shift_configs:
            try:
                h, m = map(int, sc.start_time_str.split(":"))
                shift_starts.append((sc.shift_name, h, m))
            except (ValueError, AttributeError):
                pass
    if not shift_starts:
        shift_starts = [("白班", 8, 30), ("夜班", 20, 30)]
    shift_starts.sort(key=lambda x: (x[1], x[2]))
    return shift_starts


def _get_shift_label(dt_val, shift_starts):
    """返回给定 datetime 的班次标签，如 "8/25D" 或 "8/25N" """
    from datetime import time as dt_time, timedelta as td
    from datetime import datetime
    shift_type_map = {"白班": "D", "夜班": "N"}
    if not shift_starts:
        return ""
    n_shifts = len(shift_starts)
    for day_offset in [0, -1]:
        check_date = dt_val.date() + td(days=day_offset)
        for i, (sname, sh, sm) in enumerate(shift_starts):
            shift_start = datetime.combine(check_date, dt_time(sh, sm))
            next_idx = (i + 1) % n_shifts
            next_name, next_h, next_m = shift_starts[next_idx]
            if next_idx <= i:
                shift_end = datetime.combine(check_date + td(days=1), dt_time(next_h, next_m))
            else:
                shift_end = datetime.combine(check_date, dt_time(next_h, next_m))
            if shift_start <= dt_val < shift_end:
                return shift_start.strftime("%m/%d") + shift_type_map.get(sname, sname)
    return ""


def export_to_excel(
    lot_entries: list[ScheduleEntry],
    eqp_entries: list[EqpScheduleEntry],
    qtime_alerts: list[QTimeAlert],
    output_path: str,
    lots: list = None,
    shift_configs: list = None,
    lot_order: list = None,
):
    """导出所有排程结果到 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data(ws, start_row, end_row, num_cols):
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = cell_alignment
                cell.border = thin_border

    # ── Sheet 1: Lot 排程明细 ──
    ws1 = wb.active
    ws1.title = "Lot Schedule"
    lot_headers = ["Lot Name", "Priority", "Product", "Stage", "Step Number", "Step Name",
                   "Eqp ID", "Start Time", "End Time", "Shift", "CT(min)", "Q-time Risk"]
    for col, h in enumerate(lot_headers, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(lot_headers))

    # 构建班次开始时间列表（用于 shift label 计算）
    shift_starts = _parse_shift_starts(shift_configs)

    for i, e in enumerate(lot_entries, 2):
        ws1.cell(row=i, column=1, value=e.lot_name)
        ws1.cell(row=i, column=2, value=e.priority)
        ws1.cell(row=i, column=3, value=e.product_name)
        ws1.cell(row=i, column=4, value=e.stage_name)
        ws1.cell(row=i, column=5, value=e.step_number)
        ws1.cell(row=i, column=6, value=e.step_name)
        ws1.cell(row=i, column=7, value=e.eqp_id)
        ws1.cell(row=i, column=8, value=e.start_time.strftime("%Y/%m/%d %H:%M"))
        ws1.cell(row=i, column=9, value=e.end_time.strftime("%Y/%m/%d %H:%M"))
        ws1.cell(row=i, column=10, value=_get_shift_label(e.start_time, shift_starts))
        ws1.cell(row=i, column=11, value=e.ct)
        ws1.cell(row=i, column=12, value=e.qtime_risk)
    style_data(ws1, 2, len(lot_entries) + 1, len(lot_headers))

    for col_idx, width in enumerate([14, 10, 12, 14, 14, 40, 14, 20, 20, 12, 10, 14], 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: 设备排程明细 ──
    ws2 = wb.create_sheet("Eqp Schedule")
    eqp_headers = ["Eqp ID", "Start Time", "End Time", "Lot Name", "Step Name", "Qty"]
    for col, h in enumerate(eqp_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(eqp_headers))

    for i, e in enumerate(eqp_entries, 2):
        ws2.cell(row=i, column=1, value=e.eqp_id)
        ws2.cell(row=i, column=2, value=e.start_time.strftime("%Y/%m/%d %H:%M"))
        ws2.cell(row=i, column=3, value=e.end_time.strftime("%Y/%m/%d %H:%M"))
        ws2.cell(row=i, column=4, value=e.lot_name)
        ws2.cell(row=i, column=5, value=e.step_name)
        ws2.cell(row=i, column=6, value=e.qty)
    style_data(ws2, 2, len(eqp_entries) + 1, len(eqp_headers))

    for col_idx, width in enumerate([14, 20, 20, 14, 40, 8], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 3: Q-time 告警 ──
    ws3 = wb.create_sheet("Q-time Alerts")
    qt_headers = ["Lot Name", "Q-time Rule", "Start Time", "Deadline", "Actual End",
                  "Over(min)", "Status"]
    for col, h in enumerate(qt_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(qt_headers))

    for i, a in enumerate(qtime_alerts, 2):
        ws3.cell(row=i, column=1, value=a.lot_name)
        ws3.cell(row=i, column=2, value=a.qtime_rule)
        ws3.cell(row=i, column=3, value=a.start_time.strftime("%Y/%m/%d %H:%M"))
        ws3.cell(row=i, column=4, value=a.deadline.strftime("%Y/%m/%d %H:%M"))
        ws3.cell(row=i, column=5, value=a.actual_end.strftime("%Y/%m/%d %H:%M"))
        ws3.cell(row=i, column=6, value=a.over_minutes)
        ws3.cell(row=i, column=7, value=a.status)
    style_data(ws3, 2, len(qtime_alerts) + 1, len(qt_headers))

    for col_idx, width in enumerate([14, 18, 20, 20, 20, 12, 10], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 4: Lot 排程汇总（与 Web 界面一致的班次透视表）──
    ws4 = wb.create_sheet("Lot Summary")
    shift_type_map = {"白班": "D", "夜班": "N"}

    # 构建班次时间列表（含名称）
    shift_starts = _parse_shift_starts(shift_configs)

    # 预计算每个班次的结束时间（下一个班次开始，末班次为次日首班次）
    n_shifts = len(shift_starts)
    shift_end_map = {}  # (shift_name, h, m) -> (end_h, end_m, wraps_next_day)
    for i, (sname, sh, sm) in enumerate(shift_starts):
        next_idx = (i + 1) % n_shifts
        next_name, next_h, next_m = shift_starts[next_idx]
        wraps = next_idx <= i  # 绕到次日
        shift_end_map[(sname, sh, sm)] = (next_h, next_m, wraps)

    # 确定 lot 顺序
    if not lot_order:
        lot_order = []
        seen = set()
        for e in lot_entries:
            if e.lot_name not in seen:
                lot_order.append(e.lot_name)
                seen.add(e.lot_name)

    # 收集所有班次列和每个 lot 在每个班次结束时的步骤和阶段
    all_shift_columns = []
    lot_steps_map = {}
    lot_stages_map = {}
    lot_entries_map = {}
    from datetime import time as dt_time
    for lot_name in lot_order:
        lot_entries_sorted = sorted(
            [e for e in lot_entries if e.lot_name == lot_name],
            key=lambda e: e.start_time
        )
        if not lot_entries_sorted:
            continue
        lot_entries_map[lot_name] = lot_entries_sorted

        lot_start_dt = lot_entries_sorted[0].start_time
        lot_end_dt = lot_entries_sorted[-1].end_time
        current_date = lot_start_dt.date()
        end_date = lot_end_dt.date()

        lot_steps_map[lot_name] = {}
        lot_stages_map[lot_name] = {}
        while current_date <= end_date:
            for shift_name, h, m in shift_starts:
                # 班次结束时间
                end_h, end_m, wraps = shift_end_map[(shift_name, h, m)]
                shift_end_dt = datetime.combine(
                    current_date + timedelta(days=1) if wraps else current_date,
                    dt_time(end_h, end_m)
                )
                shift_start_dt = datetime.combine(current_date, dt_time(h, m))
                if shift_end_dt <= lot_start_dt or shift_start_dt >= lot_end_dt:
                    continue

                # 列头用班次开始时间标识，值用班次结束时的 step
                shift_key = shift_start_dt.strftime("%m/%d") + " " + shift_type_map.get(shift_name, shift_name)

                active_step = "-"
                active_entry = None
                for entry in lot_entries_sorted:
                    if entry.start_time <= shift_end_dt <= entry.end_time:
                        active_step = entry.step_name
                        active_entry = entry
                        break
                    elif entry.end_time <= shift_end_dt:
                        active_step = entry.step_name
                        active_entry = entry

                lot_steps_map[lot_name][shift_key] = active_step
                lot_stages_map[lot_name][shift_key] = active_entry.stage_name if active_entry else "-"
                if shift_key not in all_shift_columns:
                    all_shift_columns.append(shift_key)
            current_date += timedelta(days=1)

    # 写表头（按日期+班次顺序排序：D在前，N在后）
    def _shift_sort_key(k):
        parts = k.split()
        if len(parts) == 2:
            m, d = parts[0].split("/")
            shift_order = 0 if parts[1] == "D" else 1
            return (int(m), int(d), shift_order)
        return (99, 99, 99)
    all_shift_columns.sort(key=_shift_sort_key)
    headers = ["Lot Name", "Stage", "Start Shift", "Finish Shift"] + all_shift_columns
    for col, h in enumerate(headers, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers))

    # 写数据行
    row = 2
    for lot_name in lot_order:
        if lot_name not in lot_steps_map:
            continue
        ws4.cell(row=row, column=1, value=lot_name)
        # Stage: 取第一个班次的阶段
        first_shift = all_shift_columns[0] if all_shift_columns else ""
        ws4.cell(row=row, column=2, value=lot_stages_map.get(lot_name, {}).get(first_shift, "-"))
        # Start Shift / Finish Shift
        sorted_entries = lot_entries_map.get(lot_name, [])
        ws4.cell(row=row, column=3, value=_get_shift_label(sorted_entries[0].start_time, shift_starts) if sorted_entries else "-")
        ws4.cell(row=row, column=4, value=_get_shift_label(sorted_entries[-1].end_time, shift_starts) if sorted_entries else "-")
        for col_idx, shift_col in enumerate(all_shift_columns, 5):
            ws4.cell(row=row, column=col_idx, value=lot_steps_map[lot_name].get(shift_col, "-"))
        row += 1
    style_data(ws4, 2, row - 1, len(headers))

    for col_idx, width in enumerate([14, 14, 12, 12] + [22] * len(all_shift_columns), 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 5: Daily Stage Summary ──
    ws5 = wb.create_sheet("Daily Stage Summary")
    # 使用 shift_configs 的第一个班次开始时间作为每日 8:30 cutoff
    cutoff_h, cutoff_m = 8, 30
    if shift_starts:
        cutoff_h, cutoff_m = shift_starts[0][1], shift_starts[0][2]

    # 收集所有日期和每个 lot 每天在 cutoff 时的 stage
    all_dates = set()
    lot_date_stages = {}  # lot_name -> {date_str -> stage_name}
    for lot_name in lot_order:
        sorted_entries = lot_entries_map.get(lot_name, [])
        if not sorted_entries:
            continue
        lot_date_stages[lot_name] = {}
        lot_start = sorted_entries[0].start_time
        lot_end = sorted_entries[-1].end_time
        current_date = lot_start.date()
        end_date = lot_end.date()
        while current_date <= end_date:
            cutoff_dt = datetime.combine(current_date, dt_time(cutoff_h, cutoff_m))
            date_str = current_date.strftime("%m/%d")
            all_dates.add(date_str)
            # 找到 cutoff 时正在执行的步骤
            active_entry = None
            for entry in sorted_entries:
                if entry.start_time <= cutoff_dt <= entry.end_time:
                    active_entry = entry
                    break
                elif entry.end_time <= cutoff_dt:
                    active_entry = entry
            lot_date_stages[lot_name][date_str] = active_entry.stage_name if active_entry else "-"
            current_date += timedelta(days=1)

    # 排序日期列
    sorted_dates = sorted(all_dates, key=lambda d: (int(d.split("/")[0]), int(d.split("/")[1])))

    headers5 = ["Lot Name"] + sorted_dates
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    style_header(ws5, 1, len(headers5))

    row = 2
    for lot_name in lot_order:
        if lot_name not in lot_date_stages:
            continue
        ws5.cell(row=row, column=1, value=lot_name)
        for col_idx, date_str in enumerate(sorted_dates, 2):
            ws5.cell(row=row, column=col_idx, value=lot_date_stages[lot_name].get(date_str, "-"))
        row += 1
    style_data(ws5, 2, row - 1, len(headers5))

    for col_idx, width in enumerate([14] + [14] * len(sorted_dates), 1):
        ws5.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    print(f"Excel exported: {output_path}")


# ============================================================
# 交互式 HTML 甘特图（纯前端 SVG + 原生 JS，无外部依赖）
# ============================================================

def plot_gantt_html(
    gantt_type: str,  # "lot" | "eqp"
    entries: list,
    time_start: Optional[datetime] = None,
    time_end: Optional[datetime] = None,
    filter_ids: Optional[list[str]] = None,
    shift_times: list[tuple[int, int]] = None,
) -> str:
    """生成自包含的交互式 HTML 甘特图（滚轮缩放 / 拖拽平移 / hover 详情）。

    返回完整 HTML 字符串，可直接内嵌 iframe 或新标签页打开。
    """
    if gantt_type == "lot":
        rows_data = {}
        color_map = _step_color_map(entries)
        for e in entries:
            if time_start and e.end_time < time_start:
                continue
            if time_end and e.start_time > time_end:
                continue
            if filter_ids and e.lot_name not in set(filter_ids):
                continue
            rows_data.setdefault(e.lot_name, []).append({
                "start": e.start_time, "end": e.end_time,
                "color": color_map.get(e.step_name, "#1f77b4"),
                "label": _short_step_name(e.step_name),
                "sub": f"{e.step_name} | {e.eqp_id} | CT={e.ct:.0f}min",
            })
    else:
        rows_data = {}
        color_map = _lot_color_map(entries)
        for e in entries:
            if time_start and e.end_time < time_start:
                continue
            if time_end and e.start_time > time_end:
                continue
            if filter_ids and e.eqp_id not in set(filter_ids):
                continue
            rows_data.setdefault(e.eqp_id, []).append({
                "start": e.start_time, "end": e.end_time,
                "color": color_map.get(e.lot_name, "#1f77b4"),
                "label": e.lot_name,
                "sub": f"{e.step_name} | qty={e.qty}",
            })

    # 确定时间范围
    all_times = [t for bars in rows_data.values() for b in bars for t in (b["start"], b["end"])]
    if not all_times:
        return "<html><body style='font-family:sans-serif;color:#999;text-align:center;padding:40px'>暂无数据</body></html>"
    g_min = min(all_times) - timedelta(hours=2)
    g_max = max(all_times) + timedelta(hours=2)

    rows = list(rows_data.keys())
    import json
    payload = {
        "rows": rows,
        "data": {r: rows_data[r] for r in rows},
        "t0": g_min.isoformat(),
        "t1": g_max.isoformat(),
        "type": gantt_type,
        "shifts": [list(t) for t in (shift_times or [])],
    }
    data_json = json.dumps(payload, ensure_ascii=False,
                           default=lambda o: o.isoformat() if isinstance(o, datetime) else str(o))

    return _GANTT_HTML_TEMPLATE.replace("{data_json}", data_json)


_GANTT_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>交互式甘特图</title>
<style>
  body { margin:0; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; background:#f5f6f8; }
  .toolbar { padding:8px 14px; background:#fff; border-bottom:1px solid #ddd; display:flex; gap:10px; align-items:center; font-size:13px; color:#555; }
  .toolbar button { padding:4px 12px; border:1px solid #ccc; border-radius:4px; background:#fff; cursor:pointer; font-size:12px; }
  .toolbar button:hover { border-color:#4096ff; color:#4096ff; }
  #ganttWrap { overflow:hidden; cursor:grab; }
  #ganttWrap.dragging { cursor:grabbing; }
  #tooltip { position:fixed; display:none; background:rgba(20,20,20,.9); color:#fff; padding:6px 10px; border-radius:6px; font-size:12px; pointer-events:none; z-index:99; max-width:320px; line-height:1.5; }
</style>
</head>
<body>
<div class="toolbar">
  <strong>交互式甘特图</strong>
  <span>滚轮缩放 | 拖拽平移 | 悬停查看详情</span>
  <span style="flex:1"></span>
  <button onclick="resetView()">重置视图</button>
</div>
<div id="ganttWrap"></div>
<div id="tooltip"></div>
<script>
const DATA = {data_json};
const rows = DATA.rows;
const bars = {};
for (const r of rows) bars[r] = DATA.data[r] || [];
let t0 = new Date(DATA.t0).getTime();
let t1 = new Date(DATA.t1).getTime();
const ROW_H = 30, LABEL_W = 150, HEAD_H = 34, BAR_H = 22;

function fmt(t) { const d=new Date(t); const p=n=>String(n).padStart(2,'0');
  return (d.getMonth()+1)+'/'+p(d.getDate())+' '+p(d.getHours())+':'+p(d.getMinutes()); }
function durLabel(ms){ const m=Math.round(ms/60000); if(m<60) return m+'min';
  const h=(m/60).toFixed(1); return h+'h'; }

function render(){
  const wrap=document.getElementById('ganttWrap');
  const W=wrap.clientWidth||960, H=HEAD_H+rows.length*ROW_H;
  // start/end 为 ISO 字符串，统一转换为毫秒时间戳用于坐标计算
  const T=(s)=>new Date(s).getTime();
  const x=(t)=>LABEL_W + (t-t0)/(t1-t0)*(W-LABEL_W);
  const span=Math.max(1,(t1-t0)/60000);
  let html=`<svg width="${W}" height="${H}" xmlns="http://www.w3.org/2000/svg">`;
  html+=`<rect x="0" y="0" width="${W}" height="${H}" fill="#fbfcfe"/>`;
  // 网格 + 时间轴刻度（约每 120px 一个刻度）
  const targetSteps=Math.max(1,Math.floor((W-LABEL_W)/120));
  const stepMin=span/targetSteps;
  const niceStep=Math.ceil(stepMin/30)*30; // 30min 粒度
  for(let ts=Math.floor(t0/60000/niceStep)*niceStep*60000; ts<=t1; ts+=niceStep*60000){
    const X=x(ts); if(X<LABEL_W) continue;
    html+=`<line x1="${X}" y1="${HEAD_H}" x2="${X}" y2="${H}" stroke="#e8e8e8" stroke-width="1"/>`;
    html+=`<text x="${X+4}" y="${HEAD_H-10}" font-size="11" fill="#888">${fmt(ts)}</text>`;
  }
  // 班次边界
  if(DATA.shifts && DATA.shifts.length){ const sd=new Date(t0); sd.setHours(0,0,0,0);
    for(let d=new Date(sd); d.getTime()<=t1+86400000; d=new Date(d.getTime()+86400000)){
      for(const [h,m] of DATA.shifts){ const st=new Date(d); st.setHours(h,m,0,0);
        if(st.getTime()>t0 && st.getTime()<t1){ const X=x(st.getTime());
          html+=`<line x1="${X}" y1="${HEAD_H}" x2="${X}" y2="${H}" stroke="#bbb" stroke-width="1" stroke-dasharray="4,4"/>`; } } } }
  // 行
  rows.forEach((r,ri)=>{
    const y=HEAD_H+ri*ROW_H;
    html+=`<text x="${LABEL_W-8}" y="${y+ROW_H/2+4}" font-size="12" fill="#333" text-anchor="end">${r}</text>`;
    html+=`<line x1="${LABEL_W}" y1="${y}" x2="${W}" y2="${y}" stroke="#f0f0f0"/>`;
    for(const b of bars[r]){
      const bxs=T(b.start), bxe=T(b.end);
      const bx=x(bxs), bw=Math.max(2, x(bxe)-bx);
      html+=`<rect class="bar" x="${bx}" y="${y+(ROW_H-BAR_H)/2}" width="${bw}" height="${BAR_H}" rx="3"
        fill="${b.color}" opacity="0.85" stroke="#fff" stroke-width="1"
        data-row="${ri}" data-title="${r}" data-bar='${JSON.stringify(b).replace(/'/g,"&#39;")}' />`;
      if(bw>40){ html+=`<text x="${bx+4}" y="${y+ROW_H/2+4}" font-size="10" fill="#fff">${b.label}</text>`; }
    }
  });
  html+=`</svg>`;
  wrap.innerHTML=html;
  // hover
  wrap.querySelectorAll('.bar').forEach(rect=>{
    rect.addEventListener('mousemove',(ev)=>{ const tip=document.getElementById('tooltip');
      const b=JSON.parse(rect.getAttribute('data-bar').replace(/&#39;/g,"'"));
      tip.style.display='block'; tip.style.left=(ev.clientX+14)+'px'; tip.style.top=(ev.clientY+14)+'px';
      tip.innerHTML=`<b>${rect.getAttribute('data-title')}</b><br>${b.sub}<br><b>${fmt(b.start)} ~ ${fmt(b.end)}</b>（${durLabel(new Date(b.end)-new Date(b.start))}）`; });
    rect.addEventListener('mouseleave',()=>{ document.getElementById('tooltip').style.display='none'; });
  });
}
function zoom(factor, cx){
  const wrap=document.getElementById('ganttWrap'); const W=wrap.clientWidth||960;
  const rel=(cx-LABEL_W)/(W-LABEL_W);
  const mid=t0+(t1-t0)*rel;
  const half=(t1-t0)/2*factor;
  t0=mid-half; t1=mid+half;
  if(t1-t0 < 3600000){ const c=(t0+t1)/2; t0=c-1800000; t1=c+1800000; }
  render();
}
let dragging=null;
const wrap=document.getElementById('ganttWrap');
wrap.addEventListener('wheel',(ev)=>{ ev.preventDefault();
  const rect=wrap.getBoundingClientRect(); zoom(ev.deltaY>0?1.4:0.7, ev.clientX-rect.left); }, {passive:false});
wrap.addEventListener('mousedown',(ev)=>{ dragging={x:ev.clientX}; wrap.classList.add('dragging'); });
window.addEventListener('mousemove',(ev)=>{ if(!dragging) return;
  const W=wrap.clientWidth||960; const dt=(ev.clientX-dragging.x)/(W-LABEL_W)*(t1-t0);
  t0-=dt; t1-=dt; dragging={x:ev.clientX}; render(); });
window.addEventListener('mouseup',()=>{ dragging=null; wrap.classList.remove('dragging'); });
function resetView(){ t0=new Date(DATA.t0).getTime(); t1=new Date(DATA.t1).getTime(); render(); }
window.addEventListener('resize', render);
render();
</script>
</body>
</html>
"""